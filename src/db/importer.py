# src/db/importer.py
import csv
import os

from src.db.item_repo import upsert_by_sku

# Excel support (optional dependency)
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def _parse_float_or_none(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        # allow commas "12,345.6"
        try:
            return float(s.replace(",", ""))
        except Exception:
            return None


def _clean_row(row: dict) -> dict:
    # normalize keys (case-insensitive headers)
    # e.g. "SKU" -> "sku"
    row = {str(k).strip().lower(): v for k, v in (row or {}).items() if k is not None}

    sku = (row.get("sku") or "").strip().upper()
    name = (row.get("name") or "").strip()

    category = (row.get("category") or "").strip().upper()

    unit_primary = (row.get("unit_primary") or "").strip().lower() or "sqft"
    unit_secondary = (row.get("unit_secondary") or "").strip().lower() or None

    sqft_per_unit = _parse_float_or_none(row.get("sqft_per_unit"))

    # ✅ new optional fields
    material = (row.get("material") or "").strip() or None
    thickness = (row.get("thickness") or "").strip() or None
    finish = (row.get("finish") or "").strip() or None

    data = {
        "sku": sku,
        "name": name,
        "category": category,
        "unit_primary": unit_primary,
        "unit_secondary": unit_secondary,
        "sqft_per_unit": sqft_per_unit,
        "material": material,
        "thickness": thickness,
        "finish": finish,
    }

    # enforce rules for BLOCK/TABLE
    if category in ("BLOCK", "TABLE"):
        data["unit_primary"] = "piece"
        data["unit_secondary"] = None
        data["sqft_per_unit"] = None

    # if no secondary unit, sqft_per_unit should be None
    if not data["unit_secondary"]:
        data["sqft_per_unit"] = None

    return data


def import_items_csv(
    db,
    file_path: str,
    mode="upsert",
    batch_size=500,
    progress_cb=None,
    stop_flag=None
):
    inserted = 0
    updated = 0
    skipped = 0
    errors = []

    def cancelled():
        return stop_flag() if stop_flag else False

    with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    total = len(rows)
    if total == 0:
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": ["CSV is empty"]}

    # ✅ Detect duplicate SKUs inside the CSV itself (case-insensitive)
    seen = set()

    for i, row in enumerate(rows, start=1):
        if cancelled():
            errors.append("Import cancelled by user.")
            break

        try:
            data = _clean_row(row)

            # validation
            if not data["sku"] or not data["name"] or not data["category"]:
                skipped += 1
                continue

            sku_key = data["sku"].upper()
            if sku_key in seen:
                skipped += 1
                errors.append(f"Row {i}: duplicate SKU in file ({data['sku']})")
                continue
            seen.add(sku_key)

            res = upsert_by_sku(db, data)  # "inserted" or "updated"
            if res == "inserted":
                inserted += 1
            else:
                updated += 1

            db.flush()

            if i % batch_size == 0:
                db.commit()

        except Exception as e:
            db.rollback()
            errors.append(f"Row {i}: {e}")

        if progress_cb:
            pct = int((i / total) * 100)
            progress_cb(
                pct,
                f"Importing {i}/{total}... Inserted: {inserted} Updated: {updated} Skipped: {skipped}"
            )

    try:
        db.commit()
    except Exception:
        db.rollback()

    if progress_cb:
        progress_cb(100, "Done ✅")

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


def import_items_xlsx(
    db,
    file_path: str,
    mode="upsert",
    batch_size=500,
    progress_cb=None,
    stop_flag=None
):
    if load_workbook is None:
        raise RuntimeError("Excel import requires 'openpyxl'. Install it: pip install openpyxl")

    inserted = 0
    updated = 0
    skipped = 0
    errors = []

    def cancelled():
        return stop_flag() if stop_flag else False

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active  # first sheet

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)

    if not header_row:
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": ["Excel is empty (no header row)."]}

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    # Expected headers (recommended):
    # sku, name, category, unit_primary, unit_secondary, sqft_per_unit, material, thickness, finish

    data_rows = []
    for r in rows_iter:
        # skip fully empty rows
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        row_dict = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            row_dict[key] = r[idx] if idx < len(r) else None
        data_rows.append(row_dict)

    total = len(data_rows)
    if total == 0:
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": ["Excel has no data rows."]}

    seen = set()

    for i, row in enumerate(data_rows, start=1):
        if cancelled():
            errors.append("Import cancelled by user.")
            break

        try:
            data = _clean_row(row)

            # validation
            if not data["sku"] or not data["name"] or not data["category"]:
                skipped += 1
                continue

            sku_key = data["sku"].upper()
            if sku_key in seen:
                skipped += 1
                errors.append(f"Row {i}: duplicate SKU in file ({data['sku']})")
                continue
            seen.add(sku_key)

            res = upsert_by_sku(db, data)
            if res == "inserted":
                inserted += 1
            else:
                updated += 1

            db.flush()

            if i % batch_size == 0:
                db.commit()

        except Exception as e:
            db.rollback()
            errors.append(f"Row {i}: {e}")

        if progress_cb:
            pct = int((i / total) * 100)
            progress_cb(
                pct,
                f"Importing {i}/{total}... Inserted: {inserted} Updated: {updated} Skipped: {skipped}"
            )

    try:
        db.commit()
    except Exception:
        db.rollback()

    if progress_cb:
        progress_cb(100, "Done ✅")

    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


def import_items_file(
    db,
    file_path: str,
    mode="upsert",
    batch_size=500,
    progress_cb=None,
    stop_flag=None
):
    ext = os.path.splitext(file_path.lower())[1]
    if ext == ".csv":
        return import_items_csv(
            db, file_path,
            mode=mode,
            batch_size=batch_size,
            progress_cb=progress_cb,
            stop_flag=stop_flag
        )
    if ext == ".xlsx":
        return import_items_xlsx(
            db, file_path,
            mode=mode,
            batch_size=batch_size,
            progress_cb=progress_cb,
            stop_flag=stop_flag
        )
    raise ValueError("Unsupported file type. Please choose .csv or .xlsx")
