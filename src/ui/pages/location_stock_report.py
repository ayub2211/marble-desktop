# src/ui/pages/location_stock_report.py
import csv
from datetime import datetime

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QComboBox, QTableWidget, QTableWidgetItem,
    QPushButton, QFileDialog, QMessageBox
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QTextDocument, QPageSize
from PySide6.QtPrintSupport import QPrinter

from src.db.session import get_db
from src.db.location_repo import get_locations
from src.db.reports_repo import location_stock_summary, location_stock_by_item


class LocationStockReportPage(QWidget):
    def __init__(self):
        super().__init__()
        layout = QVBoxLayout(self)

        # ---------------------------
        # Header + Actions
        # ---------------------------
        header = QHBoxLayout()
        title = QLabel("Location-wise Stock Report")
        title.setStyleSheet("font-size:22px;font-weight:800;")

        self.export_csv_btn = QPushButton("Export CSV")
        self.export_xlsx_btn = QPushButton("Export Excel (.xlsx)")
        self.print_pdf_btn = QPushButton("Save PDF (Landscape)")

        self.export_csv_btn.clicked.connect(self.export_csv)
        self.export_xlsx_btn.clicked.connect(self.export_xlsx)
        self.print_pdf_btn.clicked.connect(self.print_or_pdf)

        header.addWidget(title)
        header.addStretch()
        header.addWidget(self.export_csv_btn)
        header.addWidget(self.export_xlsx_btn)
        header.addWidget(self.print_pdf_btn)
        layout.addLayout(header)

        # ---------------------------
        # Filters row
        # ---------------------------
        filters = QHBoxLayout()

        self.loc_dd = QComboBox()
        self.loc_dd.addItem("All Locations", None)

        self.cat_dd = QComboBox()
        self.cat_dd.addItems(["ALL", "SLAB", "TILE", "BLOCK", "TABLE"])

        self.search = QLineEdit()
        self.search.setPlaceholderText("Search SKU / Name...")

        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.clicked.connect(self.reload)

        filters.addWidget(QLabel("Location:"))
        filters.addWidget(self.loc_dd, 1)
        filters.addSpacing(10)
        filters.addWidget(QLabel("Category:"))
        filters.addWidget(self.cat_dd, 1)
        filters.addSpacing(10)
        filters.addWidget(self.search, 2)
        filters.addWidget(self.refresh_btn)
        layout.addLayout(filters)

        # ---------------------------
        # Summary table
        # ---------------------------
        self.summary_label = QLabel("Summary (Totals by Location):")
        self.summary_label.setStyleSheet("font-weight:700; margin-top:8px;")
        layout.addWidget(self.summary_label)

        self.summary_table = QTableWidget(0, 7)
        self.summary_table.setHorizontalHeaderLabels([
            "Location", "Slabs (count)", "Slabs (sqft)", "Tiles (boxes)",
            "Tiles (sqft)", "Blocks (pieces)", "Tables (pieces)"
        ])
        self.summary_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.summary_table)

        # ---------------------------
        # Items table
        # ---------------------------
        self.items_label = QLabel("Item-wise Stock:")
        self.items_label.setStyleSheet("font-weight:700; margin-top:10px;")
        layout.addWidget(self.items_label)

        self.items_table = QTableWidget(0, 7)
        self.items_table.setHorizontalHeaderLabels([
            "Location", "Category", "SKU", "Name",
            "Primary Qty", "Primary Unit", "Secondary Qty"
        ])
        self.items_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.items_table)

        # signals
        self.loc_dd.currentIndexChanged.connect(self.reload)
        self.cat_dd.currentIndexChanged.connect(self.reload)
        self.search.textChanged.connect(self.reload)

        self._load_locations()
        self.reload()

    # ---------------------------
    # Helpers
    # ---------------------------
    def _load_locations(self):
        with get_db() as db:
            locs = get_locations(db)

        self.loc_dd.blockSignals(True)
        self.loc_dd.clear()
        self.loc_dd.addItem("All Locations", None)
        for l in locs:
            self.loc_dd.addItem(l.name, l.id)
        self.loc_dd.blockSignals(False)

    def _make_item(self, text: str, bold=False, align=Qt.AlignLeft):
        it = QTableWidgetItem(str(text if text is not None else ""))
        it.setTextAlignment(align)
        if bold:
            f = it.font()
            f.setBold(True)
            it.setFont(f)
        return it

    def reload(self):
        self._load_summary()
        self._load_items()

    # ---------------------------
    # Summary
    # ---------------------------
    def _load_summary(self):
        self.summary_table.setRowCount(0)

        with get_db() as db:
            rows = location_stock_summary(db)

        for r, row in enumerate(rows):
            self.summary_table.insertRow(r)
            self.summary_table.setItem(r, 0, self._make_item(row["location_name"]))
            self.summary_table.setItem(r, 1, self._make_item(row["slab_count"], align=Qt.AlignRight))
            self.summary_table.setItem(r, 2, self._make_item(f'{float(row["slab_sqft"]):.3f}', align=Qt.AlignRight))
            self.summary_table.setItem(r, 3, self._make_item(row["tile_boxes"], align=Qt.AlignRight))
            self.summary_table.setItem(r, 4, self._make_item(f'{float(row["tile_sqft"]):.3f}', align=Qt.AlignRight))
            self.summary_table.setItem(r, 5, self._make_item(row["block_pieces"], align=Qt.AlignRight))
            self.summary_table.setItem(r, 6, self._make_item(row["table_pieces"], align=Qt.AlignRight))

        # location filter (hide)
        loc_id = self.loc_dd.currentData()
        if loc_id is not None:
            chosen = self.loc_dd.currentText()
            for rr in range(self.summary_table.rowCount()):
                loc_name = self.summary_table.item(rr, 0).text()
                self.summary_table.setRowHidden(rr, loc_name != chosen)
        else:
            for rr in range(self.summary_table.rowCount()):
                self.summary_table.setRowHidden(rr, False)

        self._append_summary_totals_row()

    def _append_summary_totals_row(self):
        # remove old totals if present at end
        if self.summary_table.rowCount() > 0:
            last = self.summary_table.rowCount() - 1
            it = self.summary_table.item(last, 0)
            if it and it.text() == "Grand Total":
                self.summary_table.removeRow(last)

        totals = {
            "slab_count": 0, "slab_sqft": 0.0,
            "tile_boxes": 0, "tile_sqft": 0.0,
            "block_pieces": 0, "table_pieces": 0
        }

        for r in range(self.summary_table.rowCount()):
            if self.summary_table.isRowHidden(r):
                continue
            totals["slab_count"] += int(self.summary_table.item(r, 1).text() or 0)
            totals["slab_sqft"] += float(self.summary_table.item(r, 2).text() or 0)
            totals["tile_boxes"] += int(self.summary_table.item(r, 3).text() or 0)
            totals["tile_sqft"] += float(self.summary_table.item(r, 4).text() or 0)
            totals["block_pieces"] += int(self.summary_table.item(r, 5).text() or 0)
            totals["table_pieces"] += int(self.summary_table.item(r, 6).text() or 0)

        rr = self.summary_table.rowCount()
        self.summary_table.insertRow(rr)
        self.summary_table.setItem(rr, 0, self._make_item("Grand Total", bold=True))
        self.summary_table.setItem(rr, 1, self._make_item(totals["slab_count"], bold=True, align=Qt.AlignRight))
        self.summary_table.setItem(rr, 2, self._make_item(f'{totals["slab_sqft"]:.3f}', bold=True, align=Qt.AlignRight))
        self.summary_table.setItem(rr, 3, self._make_item(totals["tile_boxes"], bold=True, align=Qt.AlignRight))
        self.summary_table.setItem(rr, 4, self._make_item(f'{totals["tile_sqft"]:.3f}', bold=True, align=Qt.AlignRight))
        self.summary_table.setItem(rr, 5, self._make_item(totals["block_pieces"], bold=True, align=Qt.AlignRight))
        self.summary_table.setItem(rr, 6, self._make_item(totals["table_pieces"], bold=True, align=Qt.AlignRight))

    # ---------------------------
    # Items
    # ---------------------------
    def _load_items(self):
        self.items_table.setRowCount(0)

        loc_id = self.loc_dd.currentData()
        cat = self.cat_dd.currentText()
        q = self.search.text().strip()

        with get_db() as db:
            rows = location_stock_by_item(db=db, location_id=loc_id, category=cat, q_text=q)

        for r, row in enumerate(rows):
            self.items_table.insertRow(r)

            primary_qty = float(row.get("primary_qty") or 0)
            secondary_qty = row.get("secondary_qty", None)

            primary_unit = row.get("primary_unit") or ""
            secondary_unit = row.get("secondary_unit") or ""

            sec_txt = ""
            if secondary_qty is not None:
                sec_txt = f"{int(secondary_qty)} ({secondary_unit})" if secondary_unit else str(int(secondary_qty))

            self.items_table.setItem(r, 0, self._make_item(row.get("location_name") or ""))
            self.items_table.setItem(r, 1, self._make_item(row.get("category") or ""))
            self.items_table.setItem(r, 2, self._make_item(row.get("sku") or ""))
            self.items_table.setItem(r, 3, self._make_item(row.get("name") or ""))
            self.items_table.setItem(r, 4, self._make_item(f"{primary_qty:.3f}", align=Qt.AlignRight))
            self.items_table.setItem(r, 5, self._make_item(primary_unit))
            self.items_table.setItem(r, 6, self._make_item(sec_txt, align=Qt.AlignRight))

        self._append_items_totals_row()

    def _append_items_totals_row(self):
        if self.items_table.rowCount() > 0:
            last = self.items_table.rowCount() - 1
            it = self.items_table.item(last, 3)
            if it and it.text() == "Grand Total":
                self.items_table.removeRow(last)

        total_primary = 0.0
        total_secondary = 0

        for r in range(self.items_table.rowCount()):
            try:
                total_primary += float(self.items_table.item(r, 4).text() or 0)
            except Exception:
                pass

            try:
                sec_txt = (self.items_table.item(r, 6).text() or "").strip()
                if sec_txt:
                    only_num = sec_txt.split(" ")[0]  # "5" from "5 (slab)"
                    total_secondary += int(only_num)
            except Exception:
                pass

        rr = self.items_table.rowCount()
        self.items_table.insertRow(rr)

        self.items_table.setItem(rr, 0, self._make_item("", bold=True))
        self.items_table.setItem(rr, 1, self._make_item("", bold=True))
        self.items_table.setItem(rr, 2, self._make_item("", bold=True))
        self.items_table.setItem(rr, 3, self._make_item("Grand Total", bold=True))
        self.items_table.setItem(rr, 4, self._make_item(f"{total_primary:.3f}", bold=True, align=Qt.AlignRight))
        self.items_table.setItem(rr, 5, self._make_item("", bold=True))
        self.items_table.setItem(rr, 6, self._make_item(str(total_secondary), bold=True, align=Qt.AlignRight))

    # ---------------------------
    # Export CSV (both tables)
    # ---------------------------
    def export_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export CSV", "location_stock_report.csv", "CSV Files (*.csv)")
        if not path:
            return

        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)

                w.writerow(["Location-wise Stock Report"])
                w.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"])
                w.writerow([f"Location: {self.loc_dd.currentText()}"])
                w.writerow([f"Category: {self.cat_dd.currentText()}"])
                w.writerow([f"Search: {self.search.text().strip() or '—'}"])
                w.writerow([])

                w.writerow(["Summary (Totals by Location)"])
                self._write_table_csv(self.summary_table, w)
                w.writerow([])
                w.writerow(["Item-wise Stock"])
                self._write_table_csv(self.items_table, w)

            QMessageBox.information(self, "Exported", f"Saved ✅\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"CSV export failed:\n{e}")

    def _write_table_csv(self, table: QTableWidget, writer):
        headers = []
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            headers.append(h.text() if h else f"col_{c}")
        writer.writerow(headers)

        for r in range(table.rowCount()):
            if table.isRowHidden(r):
                continue
            row_vals = []
            for c in range(table.columnCount()):
                it = table.item(r, c)
                row_vals.append(it.text() if it else "")
            writer.writerow(row_vals)

    # ---------------------------
    # Export Excel (.xlsx)
    # ---------------------------
    def export_xlsx(self):
        path, _ = QFileDialog.getSaveFileName(self, "Export Excel", "location_stock_report.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except Exception:
            QMessageBox.critical(self, "Missing dependency", "openpyxl not installed.\n\nRun:\npip install openpyxl")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            ws.append(["Location-wise Stock Report"])
            ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"])
            ws.append([f"Location: {self.loc_dd.currentText()}"])
            ws.append([f"Category: {self.cat_dd.currentText()}"])
            ws.append([f"Search: {self.search.text().strip() or '—'}"])
            ws.append([])

            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="left")

            ws.append(["Summary (Totals by Location)"])
            ws[f"A{ws.max_row}"].font = Font(bold=True)
            self._write_table_xlsx(ws, self.summary_table)

            ws.append([])
            ws.append(["Item-wise Stock"])
            ws[f"A{ws.max_row}"].font = Font(bold=True)
            self._write_table_xlsx(ws, self.items_table)

            wb.save(path)
            QMessageBox.information(self, "Exported", f"Saved ✅\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Excel export failed:\n{e}")

    def _write_table_xlsx(self, ws, table: QTableWidget):
        from openpyxl.styles import Font, Alignment
        bold = Font(bold=True)

        headers = []
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            headers.append(h.text() if h else f"col_{c}")

        ws.append(headers)
        header_row = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")

        for r in range(table.rowCount()):
            if table.isRowHidden(r):
                continue
            row_vals = []
            for c in range(table.columnCount()):
                it = table.item(r, c)
                row_vals.append(it.text() if it else "")
            ws.append(row_vals)

        # bold totals row if "Grand Total" exists
        for rr in range(ws.max_row, max(ws.max_row - 4, 1), -1):
            if any(str(ws.cell(row=rr, column=cc).value).strip() == "Grand Total" for cc in range(1, ws.max_column + 1)):
                for cc in range(1, ws.max_column + 1):
                    ws.cell(row=rr, column=cc).font = bold
                break

    # ---------------------------
    # Save PDF (A4 Landscape, FULL REPORT)
    # ---------------------------
    def print_or_pdf(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save PDF", "location_stock_report.pdf", "PDF Files (*.pdf)")
        if not path:
            return

        filters = {
            "Location": self.loc_dd.currentText(),
            "Category": self.cat_dd.currentText(),
            "Search": self.search.text().strip() or "—",
        }
        html = self._full_report_html(filters)

        doc = QTextDocument()
        doc.setHtml(html)

        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(path)

        # ✅ A4 Landscape
        from PySide6.QtGui import QPageLayout
        page_size = QPageSize(QPageSize.A4)
        layout = QPageLayout(page_size, QPageLayout.Landscape, QPageLayout.MarginsF(12, 12, 12, 12))
        printer.setPageLayout(layout)

        doc.setPageSize(printer.pageRect().size())
        doc.print_(printer)

        QMessageBox.information(self, "Saved", f"PDF saved ✅\n{path}")

    # ---------------------------
    # HTML builders
    # ---------------------------
    def _full_report_html(self, filters=None) -> str:
        filters = filters or {}
        now = datetime.now().strftime("%Y-%m-%d %I:%M %p")

        flt_html = ""
        if filters:
            flt_rows = "".join([
                f"<div><b>{self._escape_html(k)}:</b> {self._escape_html(v)}</div>"
                for k, v in filters.items()
            ])
            flt_html = f'<div class="filters">{flt_rows}</div>'

        summary_html = self._table_html(self.summary_table)
        items_html = self._table_html(self.items_table)

        return f"""
        <html>
        <head>
        <meta charset="utf-8" />
        <style>
            body {{
                font-family: Arial, sans-serif;
                font-size: 10.5pt;
                color: #111;
            }}
            h1 {{
                margin: 0 0 6px 0;
                font-size: 18pt;
            }}
            .meta {{
                color: #444;
                font-size: 9.5pt;
                margin-bottom: 8px;
            }}
            .filters {{
                font-size: 9.8pt;
                color: #222;
                margin-bottom: 12px;
            }}
            .filters div {{
                margin: 2px 0;
            }}
            h2 {{
                margin: 14px 0 6px 0;
                font-size: 12.5pt;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                table-layout: fixed;
            }}
            th, td {{
                border: 1px solid #999;
                padding: 6px;
                vertical-align: top;
                word-wrap: break-word;
            }}
            th {{
                background: #f2f2f2;
                font-weight: bold;
            }}
            .totals td {{
                font-weight: bold;
                background: #fafafa;
            }}
        </style>
        </head>
        <body>
            <h1>Location-wise Stock Report</h1>
            <div class="meta">Generated: {self._escape_html(now)}</div>
            {flt_html}

            <h2>Summary (Totals by Location)</h2>
            {summary_html}

            <h2>Item-wise Stock</h2>
            {items_html}
        </body>
        </html>
        """

    def _table_html(self, table: QTableWidget) -> str:
        headers = []
        for c in range(table.columnCount()):
            h = table.horizontalHeaderItem(c)
            headers.append(h.text() if h else f"col_{c}")

        ths = "".join([f"<th>{self._escape_html(h)}</th>" for h in headers])

        trs = ""
        for r in range(table.rowCount()):
            if table.isRowHidden(r):
                continue

            is_total = False
            for c in range(table.columnCount()):
                it = table.item(r, c)
                if it and it.text().strip() == "Grand Total":
                    is_total = True
                    break

            row_vals = []
            for c in range(table.columnCount()):
                it = table.item(r, c)
                row_vals.append(it.text() if it else "")

            tds = "".join([f"<td>{self._escape_html(v)}</td>" for v in row_vals])
            trs += f'<tr class="totals">{tds}</tr>' if is_total else f"<tr>{tds}</tr>"

        return f"""
        <table>
            <thead><tr>{ths}</tr></thead>
            <tbody>{trs}</tbody>
        </table>
        """

    def _escape_html(self, s: str) -> str:
        s = str(s or "")
        return (
            s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace(">", "&gt;")
             .replace('"', "&quot;")
        )
